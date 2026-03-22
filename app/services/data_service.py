"""
数据备份、导出和导入服务
"""
import io
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.models import Student, Teacher, Course, Enrollment, db
from app.utils import FileOperationError, log_audit


class BackupService:
    """数据备份服务"""
    
    BACKUP_DIR = Path("backups")
    
    @staticmethod
    def ensure_backup_dir():
        """确保备份目录存在"""
        BackupService.BACKUP_DIR.mkdir(exist_ok=True)
    
    @staticmethod
    def create_backup(backup_name=None):
        """
        创建数据库备份
        
        Args:
            backup_name: 备份名称（不指定则自动生成）
        
        Returns:
            备份文件路径
        """
        BackupService.ensure_backup_dir()
        
        if not backup_name:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            backup_name = f"backup_{timestamp}"
        
        backup_path = BackupService.BACKUP_DIR / f"{backup_name}.db"
        
        try:
            # 复制数据库文件
            source_db = Path("students.db")
            if not source_db.exists():
                raise FileOperationError("源数据库文件不存在")
            
            shutil.copy2(source_db, backup_path)
            
            return str(backup_path)
        except Exception as e:
            raise FileOperationError(f"创建备份失败: {str(e)}")
    
    @staticmethod
    def restore_backup(backup_path):
        """
        恢复数据库备份
        
        Args:
            backup_path: 备份文件路径
        """
        backup_file = Path(backup_path)
        if not backup_file.exists():
            raise FileOperationError("备份文件不存在")
        
        target_db = Path("students.db")
        
        try:
            # 创建当前数据库的备份（防止丢失）
            if target_db.exists():
                temp_backup = Path(f"students_temp_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.db")
                shutil.copy2(target_db, temp_backup)
            
            # 恢复备份
            shutil.copy2(backup_file, target_db)
            
            # 重新初始化数据库连接
            db.session.close()
            
            return True
        except Exception as e:
            raise FileOperationError(f"恢复备份失败: {str(e)}")
    
    @staticmethod
    def list_backups():
        """列出所有备份文件"""
        BackupService.ensure_backup_dir()
        backups = []
        for backup_file in BackupService.BACKUP_DIR.glob("*.db"):
            stat = backup_file.stat()
            backups.append({
                'name': backup_file.stem,
                'path': str(backup_file),
                'size': stat.st_size,
                'created_at': datetime.fromtimestamp(stat.st_mtime)
            })
        return sorted(backups, key=lambda x: x['created_at'], reverse=True)
    
    @staticmethod
    def cleanup_old_backups(days=30):
        """清理旧备份"""
        from datetime import timedelta
        
        BackupService.ensure_backup_dir()
        cutoff_time = datetime.utcnow() - timedelta(days=days)
        
        deleted_count = 0
        for backup_file in BackupService.BACKUP_DIR.glob("*.db"):
            file_time = datetime.fromtimestamp(backup_file.stat().st_mtime)
            if file_time < cutoff_time:
                try:
                    backup_file.unlink()
                    deleted_count += 1
                except Exception as e:
                    pass
        
        return deleted_count


class ExportService:
    """数据导出服务"""
    
    @staticmethod
    def _create_styled_workbook(title):
        """创建格式化的工作簿"""
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        return wb, ws, header_fill, header_font
    
    @staticmethod
    def export_students(student_ids=None):
        """
        导出学生数据
        
        Args:
            student_ids: 要导出的学生ID列表（None表示全部）
        
        Returns:
            Excel文件BytesIO对象
        """
        try:
            wb, ws, header_fill, header_font = ExportService._create_styled_workbook("学生数据")
            
            # 设置列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15
            
            # 写入表头
            headers = ['学号', '姓名', '性别', '年龄', '班级', '手机', '邮箱', '专业', '身份证', '学籍状态']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 获取学生数据
            query = Student.query
            if student_ids:
                query = query.filter(Student.id.in_(student_ids))
            students = query.all()
            
            # 写入数据
            for row_idx, student in enumerate(students, 2):
                ws.cell(row=row_idx, column=1, value=student.student_no)
                ws.cell(row=row_idx, column=2, value=student.name)
                ws.cell(row=row_idx, column=3, value=student.gender or "")
                ws.cell(row=row_idx, column=4, value=student.age or "")
                ws.cell(row=row_idx, column=5, value=student.class_obj.class_name if student.class_obj else "")
                ws.cell(row=row_idx, column=6, value=student.phone or "")
                ws.cell(row=row_idx, column=7, value=student.email or "")
                ws.cell(row=row_idx, column=8, value=student.major or "")
                ws.cell(row=row_idx, column=9, value=student.id_number or "")
                ws.cell(row=row_idx, column=10, value=student.status or "")
            
            # 输出到字节流
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            return buf
        except Exception as e:
            raise FileOperationError(f"导出学生数据失败: {str(e)}")
    
    @staticmethod
    def export_courses(course_ids=None):
        """导出课程数据"""
        try:
            wb, ws, header_fill, header_font = ExportService._create_styled_workbook("课程数据")
            
            # 设置列宽
            column_widths = [12, 20, 8, 8, 12, 12, 8, 8]
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = width
            
            # 写入表头
            headers = ['课程号', '课程名', '学分', '学时', '教师', '容量', '学期', '状态']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 获取课程数据
            query = Course.query
            if course_ids:
                query = query.filter(Course.id.in_(course_ids))
            courses = query.all()
            
            # 写入数据
            for row_idx, course in enumerate(courses, 2):
                ws.cell(row=row_idx, column=1, value=course.course_no)
                ws.cell(row=row_idx, column=2, value=course.course_name)
                ws.cell(row=row_idx, column=3, value=course.credits or "")
                ws.cell(row=row_idx, column=4, value=course.hours or "")
                ws.cell(row=row_idx, column=5, value=course.teacher.name if course.teacher else "")
                ws.cell(row=row_idx, column=6, value=course.max_capacity or "")
                ws.cell(row=row_idx, column=7, value=course.semester or "")
                ws.cell(row=row_idx, column=8, value=course.status or "")
            
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            return buf
        except Exception as e:
            raise FileOperationError(f"导出课程数据失败: {str(e)}")


class ImportService:
    """数据导入服务"""
    
    @staticmethod
    def import_students_from_excel(file_stream, created_by_id=None):
        """
        从Excel导入学生数据
        
        Returns:
            (成功数, 失败数, 错误列表)
        """
        from app.utils import DuplicateResourceError, ValidationError
        from app.services import StudentService
        from app.models import Class
        
        try:
            wb = load_workbook(file_stream, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            raise FileOperationError(f"Excel文件格式不正确: {str(e)}")
        
        success, failed, errors = 0, 0, []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v for v in row if v not in (None, "")):
                continue
            
            try:
                def _s(v):
                    return str(v).strip() if v is not None else ""
                
                student_no = _s(row[0])
                name = _s(row[1])
                gender = _s(row[2])
                age_raw = row[3]
                class_no = _s(row[4])
                phone = _s(row[5])
                email = _s(row[6])
                major = _s(row[7])
                id_number = _s(row[8])
                
                if not student_no or not name:
                    errors.append(f"第 {row_idx} 行：学号和姓名不能为空")
                    failed += 1
                    continue
                
                # 获取班级ID
                class_obj = None
                if class_no:
                    class_obj = Class.query.filter_by(class_no=class_no).first()
                    if not class_obj:
                        errors.append(f"第 {row_idx} 行：班级 {class_no} 不存在")
                        failed += 1
                        continue
                
                # 处理年龄
                age = None
                if age_raw not in (None, ""):
                    try:
                        age = int(float(str(age_raw)))
                    except (ValueError, TypeError):
                        age = None
                
                # 创建学生
                StudentService.create_student(
                    student_no=student_no,
                    name=name,
                    class_id=class_obj.id if class_obj else None,
                    gender=gender,
                    age=age,
                    phone=phone,
                    email=email,
                    major=major,
                    id_number=id_number,
                    created_by_id=created_by_id
                )
                success += 1
            except (ValidationError, DuplicateResourceError) as e:
                errors.append(f"第 {row_idx} 行：{str(e)}")
                failed += 1
            except Exception as e:
                errors.append(f"第 {row_idx} 行：导入失败 - {str(e)}")
                failed += 1
        
        return success, failed, errors[:10]  # 只返回前10个错误
