import io
import importlib.util
import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook


ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
APP_FILE = os.path.join(ROOT_DIR, "app.py")
APP_SPEC = importlib.util.spec_from_file_location("tested_app", APP_FILE)
app_module = importlib.util.module_from_spec(APP_SPEC)
APP_SPEC.loader.exec_module(app_module)


class FunctionalFlowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls._temp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        cls._temp_db.close()

        app_module.app.config.update(
            TESTING=True,
            WTF_CSRF_ENABLED=False,
            SQLALCHEMY_DATABASE_URI=f"sqlite:///{cls._temp_db.name}",
        )

        cls.app = app_module.app
        cls.db = app_module.db

    @classmethod
    def tearDownClass(cls):
        try:
            os.unlink(cls._temp_db.name)
        except OSError:
            pass

    def setUp(self):
        with self.app.app_context():
            self.db.session.remove()
            self.db.drop_all()
            self.db.create_all()
            app_module.init_db()

        self.client = self.app.test_client()

    def tearDown(self):
        with self.app.app_context():
            self.db.session.remove()

    def login(self, username, password, follow_redirects=True):
        return self.client.post(
            "/login",
            data={"username": username, "password": password},
            follow_redirects=follow_redirects,
        )

    def logout(self):
        return self.client.get("/logout", follow_redirects=True)

    def create_teacher(self, teacher_no="T001", name="张老师"):
        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)
        response = self.client.post(
            "/teachers/add",
            data={
                "teacher_no": teacher_no,
                "name": name,
                "gender": "男",
                "department": "计算机学院",
                "phone": "13800138000",
                "email": f"{teacher_no.lower()}@school.com",
                "qualification": "硕士",
            },
            follow_redirects=True,
        )
        self.assertIn("新增教师成功", response.get_data(as_text=True))

        with self.app.app_context():
            teacher = app_module.Teacher.query.filter_by(teacher_no=teacher_no).first()
            self.assertIsNotNone(teacher)
            return teacher.id, teacher.user_id

    def create_class(self, class_no="C001", class_name="软件1班", headteacher_teacher_id=None):
        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)
        response = self.client.post(
            "/classes/add",
            data={
                "class_no": class_no,
                "class_name": class_name,
                "headteacher_id": str(headteacher_teacher_id) if headteacher_teacher_id else "",
                "grade": "2024",
            },
            follow_redirects=True,
        )
        self.assertIn("新增班级成功", response.get_data(as_text=True))

        with self.app.app_context():
            cls = app_module.Class.query.filter_by(class_no=class_no).first()
            self.assertIsNotNone(cls)
            return cls.id

    def create_student(self, student_no="20240001", name="张三", class_id=None):
        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)
        response = self.client.post(
            "/students/add",
            data={
                "student_no": student_no,
                "name": name,
                "gender": "男",
                "age": "20",
                "class_id": str(class_id) if class_id else "",
                "status": "enrolled",
                "phone": "13900000000",
                "email": f"{student_no}@school.com",
                "major": "软件工程",
                "id_number": f"ID{student_no}",
            },
            follow_redirects=True,
        )
        self.assertIn("新增学生成功", response.get_data(as_text=True))

        with self.app.app_context():
            student = app_module.Student.query.filter_by(student_no=student_no).first()
            self.assertIsNotNone(student)
            return student.id, student.user_id

    def create_course(
        self,
        course_no="CS101",
        course_name="程序设计",
        teacher_id=None,
        max_capacity=50,
        semester="2026-1",
    ):
        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)
        response = self.client.post(
            "/courses/add",
            data={
                "course_no": course_no,
                "course_name": course_name,
                "credits": "3",
                "hours": "48",
                "teacher_id": str(teacher_id) if teacher_id else "",
                "co_teacher_ids": [],
                "max_capacity": str(max_capacity),
                "semester": semester,
            },
            follow_redirects=True,
        )
        self.assertIn("新增课程成功", response.get_data(as_text=True))

        with self.app.app_context():
            course = app_module.Course.query.filter_by(course_no=course_no).first()
            self.assertIsNotNone(course)
            return course.id

    def get_enrollment(self, student_id, course_id):
        with self.app.app_context():
            return app_module.Enrollment.query.filter_by(student_id=student_id, course_id=course_id).first()

    def workbook_rows(self, response):
        workbook = load_workbook(io.BytesIO(response.data), read_only=True, data_only=True)
        sheet = workbook.active
        return list(sheet.iter_rows(values_only=True))

    def test_unauthenticated_access_redirects_to_login(self):
        response = self.client.get("/scores", follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("请先登录", response.get_data(as_text=True))
        self.assertIn("登录", response.get_data(as_text=True))

    def test_admin_can_create_edit_and_delete_student_with_account_sync(self):
        teacher_id, _ = self.create_teacher(teacher_no="T100", name="辅导员老师")
        class_id = self.create_class(class_no="C100", class_name="测试班", headteacher_teacher_id=teacher_id)
        student_id, user_id = self.create_student(student_no="20241001", name="原名", class_id=class_id)

        response = self.client.post(
            f"/students/edit/{student_id}",
            data={
                "student_no": "20241099",
                "name": "新名字",
                "gender": "女",
                "age": "21",
                "class_id": str(class_id),
                "status": "leave",
                "phone": "13700000000",
                "email": "updated@school.com",
                "major": "计算机科学",
                "id_number": "ID20241099",
            },
            follow_redirects=True,
        )
        self.assertIn("修改成功，学生账号信息已同步", response.get_data(as_text=True))

        with self.app.app_context():
            student = app_module.Student.query.get(student_id)
            user = app_module.User.query.get(user_id)
            self.assertEqual(student.student_no, "20241099")
            self.assertEqual(student.name, "新名字")
            self.assertEqual(user.username, "20241099")
            self.assertEqual(user.full_name, "新名字")

        response = self.client.post(f"/students/delete/{student_id}", follow_redirects=True)
        self.assertIn("删除成功", response.get_data(as_text=True))

        with self.app.app_context():
            self.assertIsNone(app_module.Student.query.get(student_id))
            self.assertIsNone(app_module.User.query.get(user_id))

    def test_teacher_and_class_management_bind_counselor_account(self):
        teacher_id, teacher_user_id = self.create_teacher(teacher_no="T200", name="李老师")
        class_id = self.create_class(class_no="C200", class_name="网络工程1班", headteacher_teacher_id=teacher_id)

        with self.app.app_context():
            cls = app_module.Class.query.get(class_id)
            self.assertEqual(cls.headteacher_id, teacher_user_id)

    def test_enrollment_blocks_duplicate_and_capacity_overflow(self):
        teacher_id, _ = self.create_teacher(teacher_no="T300", name="授课老师")
        class_id = self.create_class(class_no="C300", class_name="物联网1班", headteacher_teacher_id=teacher_id)
        course_id = self.create_course(course_no="CS300", course_name="数据库", teacher_id=teacher_id, max_capacity=1)
        student1_id, _ = self.create_student(student_no="20243001", name="学生甲", class_id=class_id)
        student2_id, _ = self.create_student(student_no="20243002", name="学生乙", class_id=class_id)

        self.logout()
        self.login("20243001", "123456")
        response = self.client.post(f"/enrollments/add/{course_id}", follow_redirects=True)
        self.assertIn("选课成功", response.get_data(as_text=True))

        response = self.client.post(f"/enrollments/add/{course_id}", follow_redirects=True)
        self.assertIn("该学生已经选过这门课程", response.get_data(as_text=True))

        self.logout()
        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)
        response = self.client.post(
            "/enrollments/admin-add",
            data={"student_id": str(student2_id), "course_id": str(course_id)},
            follow_redirects=True,
        )
        self.assertIn("课程已满员", response.get_data(as_text=True))

        with self.app.app_context():
            self.assertIsNotNone(app_module.Enrollment.query.filter_by(student_id=student1_id, course_id=course_id).first())
            self.assertIsNone(app_module.Enrollment.query.filter_by(student_id=student2_id, course_id=course_id).first())

    def test_negative_score_is_rejected_and_student_view_keeps_expected_score(self):
        teacher_id, _ = self.create_teacher(teacher_no="T400", name="成绩老师")
        class_id = self.create_class(class_no="C400", class_name="自动化1班", headteacher_teacher_id=teacher_id)
        course_id = self.create_course(course_no="CS400", course_name="操作系统", teacher_id=teacher_id)
        student_id, _ = self.create_student(student_no="20244001", name="学生丙", class_id=class_id)

        self.logout()
        self.login("20244001", "123456")
        self.client.post(f"/enrollments/add/{course_id}", follow_redirects=True)

        enrollment = self.get_enrollment(student_id, course_id)
        self.logout()
        self.login("T400", "123456")

        response = self.client.post(
            f"/scores/record/{enrollment.id}",
            data={"score": "-1", "grade": "F"},
            follow_redirects=True,
        )
        self.assertIn("成绩必须在 0-100 之间", response.get_data(as_text=True))

        response = self.client.post(
            f"/scores/record/{enrollment.id}",
            data={"score": "80", "grade": "B"},
            follow_redirects=True,
        )
        self.assertIn("成绩录入成功", response.get_data(as_text=True))

        self.logout()
        self.login("20244001", "123456")
        response = self.client.get("/scores", follow_redirects=True)
        page = response.get_data(as_text=True)
        self.assertIn("操作系统", page)
        self.assertIn("80", page)

        with self.app.app_context():
            enrollment = app_module.Enrollment.query.get(enrollment.id)
            self.assertEqual(enrollment.score, 80)
            self.assertEqual(enrollment.grade, "B")

    def test_password_reset_key_flow_allows_login_with_new_password(self):
        self.create_student(student_no="20245001", name="找回密码学生")
        self.logout()

        response = self.client.post(
            "/forgot-password",
            data={"identifier": "20245001"},
            follow_redirects=True,
        )
        self.assertIn("申请已提交，请联系系统管理员获取重置Key", response.get_data(as_text=True))

        with self.app.app_context():
            reset_key = (
                app_module.PasswordResetKey.query
                .join(app_module.User, app_module.PasswordResetKey.user_id == app_module.User.id)
                .filter(app_module.User.username == "20245001")
                .order_by(app_module.PasswordResetKey.created_at.desc())
                .first()
            )
            self.assertIsNotNone(reset_key)
            key_value = reset_key.reset_key

        response = self.client.post(
            "/reset-password-with-key",
            data={
                "account": "20245001",
                "key": key_value,
                "new_password": "654321",
                "confirm_password": "654321",
            },
            follow_redirects=True,
        )
        self.assertIn("密码重置成功，请使用新密码登录", response.get_data(as_text=True))

        response = self.login("20245001", "654321")
        self.assertIn("欢迎 找回密码学生", response.get_data(as_text=True))

    def test_admin_cannot_delete_current_user(self):
        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)

        with self.app.app_context():
            admin_user = app_module.User.query.filter_by(username=app_module.DEFAULT_ADMIN_USERNAME).first()
            admin_id = admin_user.id

        response = self.client.post(f"/users/delete/{admin_id}", follow_redirects=True)
        self.assertIn("不能删除当前登录账号", response.get_data(as_text=True))

    def test_student_export_should_generate_workbook(self):
        teacher_id, _ = self.create_teacher(teacher_no="T500", name="导出老师")
        class_id = self.create_class(class_no="C500", class_name="大数据1班", headteacher_teacher_id=teacher_id)
        self.create_student(student_no="20246001", name="导出学生", class_id=class_id)

        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)
        response = self.client.post("/admin/data/export/students")

        self.assertEqual(response.status_code, 200)
        rows = self.workbook_rows(response)
        self.assertGreaterEqual(len(rows), 2)
        self.assertEqual(rows[1][0], "20246001")

    def test_score_export_should_include_recorded_score(self):
        teacher_id, _ = self.create_teacher(teacher_no="T600", name="导出成绩老师")
        class_id = self.create_class(class_no="C600", class_name="人工智能1班", headteacher_teacher_id=teacher_id)
        course_id = self.create_course(course_no="CS600", course_name="机器学习", teacher_id=teacher_id)
        student_id, _ = self.create_student(student_no="20247001", name="成绩导出学生", class_id=class_id)

        self.logout()
        self.login("20247001", "123456")
        self.client.post(f"/enrollments/add/{course_id}", follow_redirects=True)

        enrollment = self.get_enrollment(student_id, course_id)
        self.logout()
        self.login("T600", "123456")
        self.client.post(
            f"/scores/record/{enrollment.id}",
            data={"score": "88", "grade": "B"},
            follow_redirects=True,
        )

        self.logout()
        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)
        response = self.client.post("/admin/data/export/scores")

        self.assertEqual(response.status_code, 200)
        rows = self.workbook_rows(response)
        self.assertGreaterEqual(len(rows), 2)
        self.assertEqual(rows[1][0], "20247001")
        self.assertEqual(rows[1][4], 88)  # 成绩列：学号, 姓名, 课程号, 课程名称, 成绩...

    def test_student_import_maps_columns_by_header_when_enrollment_date_exists(self):
        teacher_id, _ = self.create_teacher(teacher_no="T700", name="导入班主任")
        self.create_class(class_no="C700", class_name="导入测试班", headteacher_teacher_id=teacher_id)

        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)

        wb = Workbook()
        ws = wb.active
        ws.title = "学生导入"
        ws.append([
            "学号*",
            "姓名*",
            "性别",
            "年龄",
            "班级编号",
            "手机",
            "邮箱",
            "专业",
            "身份证号",
            "入学日期",
            "账号",
            "密码",
        ])
        ws.append([
            "20248001",
            "错位修复学生",
            "男",
            20,
            "C700",
            "13600000000",
            "aligned@school.com",
            "软件工程",
            "ID20248001",
            "2025-09-01",
            "stu_aligned",
            "abc12345",
        ])

        payload = io.BytesIO()
        wb.save(payload)
        payload.seek(0)

        response = self.client.post(
            "/students/import",
            data={"file": (payload, "students_with_enrollment_date.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertIn("导入完成：成功 1 条", response.get_data(as_text=True))

        self.logout()
        response = self.login("stu_aligned", "abc12345")
        self.assertIn("欢迎 错位修复学生", response.get_data(as_text=True))

    def test_student_import_accepts_class_name_without_class_no(self):
        teacher_id, _ = self.create_teacher(teacher_no="T701", name="导入班主任2")
        class_id = self.create_class(class_no="C701", class_name="仅名称匹配班", headteacher_teacher_id=teacher_id)

        self.login(app_module.DEFAULT_ADMIN_USERNAME, app_module.DEFAULT_ADMIN_PASSWORD)

        wb = Workbook()
        ws = wb.active
        ws.title = "学生导入"
        ws.append([
            "学号*",
            "姓名*",
            "性别",
            "年龄",
            "班级名称",
            "手机",
            "邮箱",
            "专业",
            "账号",
            "密码",
        ])
        ws.append([
            "20248002",
            "名称导入学生",
            "女",
            19,
            "仅名称匹配班",
            "13500000000",
            "class_name_only@school.com",
            "网络工程",
            "name_only_user",
            "xyz12345",
        ])

        payload = io.BytesIO()
        wb.save(payload)
        payload.seek(0)

        response = self.client.post(
            "/students/import",
            data={"file": (payload, "students_class_name_only.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertIn("导入完成：成功 1 条", response.get_data(as_text=True))

        with self.app.app_context():
            student = app_module.Student.query.filter_by(student_no="20248002").first()
            self.assertIsNotNone(student)
            self.assertEqual(student.class_id, class_id)

        self.logout()
        response = self.login("name_only_user", "xyz12345")
        self.assertIn("欢迎 名称导入学生", response.get_data(as_text=True))


if __name__ == "__main__":
    unittest.main(verbosity=2)