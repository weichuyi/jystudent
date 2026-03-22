import os
import shutil
import platform
import secrets
from datetime import datetime, timedelta
from functools import wraps

import io

from flask import Flask, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import check_password_hash, generate_password_hash

app = Flask(__name__)
app.config["SECRET_KEY"] = "student-system-secret-key-2026"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///students.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

course_teacher_links = db.Table(
    "course_teacher_links",
    db.Column("course_id", db.Integer, db.ForeignKey("courses.id"), primary_key=True),
    db.Column("teacher_id", db.Integer, db.ForeignKey("teachers.id"), primary_key=True),
)


# ======================== 模型定义 ========================

class User(db.Model):
    """用户表：系统用户（管理员、教师、学生）"""
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="student")  # admin, teacher, student
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def set_password(self, password):
        self.password = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password, password)


class Class(db.Model):
    """班级表"""
    __tablename__ = "classes"
    id = db.Column(db.Integer, primary_key=True)
    class_no = db.Column(db.String(50), unique=True, nullable=False)
    class_name = db.Column(db.String(100), nullable=False)
    headteacher_id = db.Column(db.Integer, db.ForeignKey("users.id"))
    grade = db.Column(db.String(20))  # 年级
    total_students = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    students = db.relationship("Student", backref="class_obj", lazy=True)
    headteacher = db.relationship("User", backref="classes_managed")


class Student(db.Model):
    """学生表：扩展版本"""
    __tablename__ = "students"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), unique=True)
    student_no = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    gender = db.Column(db.String(10))
    age = db.Column(db.Integer)
    class_id = db.Column(db.Integer, db.ForeignKey("classes.id"))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    major = db.Column(db.String(100))
    status = db.Column(db.String(20), default="enrolled")  # enrolled, leave, return, dropout
    id_number = db.Column(db.String(50), unique=True)
    enrollment_date = db.Column(db.Date)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    user = db.relationship("User", backref="student_profile")
    enrollments = db.relationship("Enrollment", backref="student", lazy=True)
    status_changes = db.relationship("StudentStatusChange", backref="student", lazy=True)


class Teacher(db.Model):
    """教师表"""
    __tablename__ = "teachers"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), unique=True)
    teacher_no = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    gender = db.Column(db.String(10))
    department = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    qualification = db.Column(db.String(100))  # 学位
    hire_date = db.Column(db.Date)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    user = db.relationship("User", backref="teacher_profile")
    courses = db.relationship("Course", backref="teacher", lazy=True)
    co_courses = db.relationship("Course", secondary=course_teacher_links, back_populates="co_teachers")


class Course(db.Model):
    """课程表"""
    __tablename__ = "courses"
    id = db.Column(db.Integer, primary_key=True)
    course_no = db.Column(db.String(50), unique=True, nullable=False)
    course_name = db.Column(db.String(100), nullable=False)
    credits = db.Column(db.Float)  # 学分
    hours = db.Column(db.Integer)  # 学时
    teacher_id = db.Column(db.Integer, db.ForeignKey("teachers.id"))
    max_capacity = db.Column(db.Integer, default=50)
    semester = db.Column(db.String(20))  # 学期 如 2024-1
    status = db.Column(db.String(20), default="open")  # open, closed
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    enrollments = db.relationship("Enrollment", backref="course", lazy=True)
    co_teachers = db.relationship("Teacher", secondary=course_teacher_links, back_populates="co_courses")


class Enrollment(db.Model):
    """选课表：学生选课记录"""
    __tablename__ = "enrollments"
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("students.id"), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey("courses.id"), nullable=False)
    score = db.Column(db.Float)
    grade = db.Column(db.String(5))  # 等级 A, B, C, D, F
    status = db.Column(db.String(20), default="enrolled")  # enrolled, withdrawn, completed
    enrollment_date = db.Column(db.DateTime, default=datetime.utcnow)
    completion_date = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    __table_args__ = (db.UniqueConstraint("student_id", "course_id", name="unique_student_course"),)


class Score(db.Model):
    """成绩表：详细成绩记录"""
    __tablename__ = "scores"
    id = db.Column(db.Integer, primary_key=True)
    enrollment_id = db.Column(db.Integer, db.ForeignKey("enrollments.id"), nullable=False)
    score_value = db.Column(db.Float)
    grade = db.Column(db.String(5))
    recorded_by = db.Column(db.Integer, db.ForeignKey("users.id"))  # 录入教师
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class StudentStatusChange(db.Model):
    """学籍异动表：记录学生休学、复学、退学等"""
    __tablename__ = "student_status_changes"
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("students.id"), nullable=False)
    change_type = db.Column(db.String(20), nullable=False)  # leave, return, dropout
    reason = db.Column(db.String(255))
    approval_date = db.Column(db.Date)
    created_by = db.Column(db.Integer, db.ForeignKey("users.id"))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class OperationLog(db.Model):
    """操作日志表"""
    __tablename__ = "operation_logs"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"))
    action = db.Column(db.String(100), nullable=False)
    module = db.Column(db.String(50))  # 模块名 student, course, score
    record_id = db.Column(db.Integer)
    details = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class PasswordResetKey(db.Model):
    """密码重置Key：用于忘记密码场景"""
    __tablename__ = "password_reset_keys"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    reset_key = db.Column(db.String(32), nullable=False)
    requested_identifier = db.Column(db.String(50))
    is_used = db.Column(db.Boolean, default=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    used_at = db.Column(db.DateTime)
    request_ip = db.Column(db.String(45))


# ======================== 权限装饰器 ========================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("请先登录", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if "user_id" not in session:
                flash("请先登录", "warning")
                return redirect(url_for("login"))
            user = User.query.get(session.get("user_id"))
            if user.role not in roles:
                flash("权限不足", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ======================== 数据库初始化 ========================

def init_db():
    with app.app_context():
        db.create_all()
        # 创建或同步默认管理员账户密码
        admin = User.query.filter_by(username="admin").first()
        if not admin:
            admin = User(
                username="admin",
                full_name="系统管理员",
                role="admin",
                email="admin@school.com"
            )
            db.session.add(admin)

        # 按需求统一默认管理员密码
        admin.set_password("weichuy1")
        db.session.commit()


# ======================== 认证路由 ========================

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password) and user.is_active:
            session["user_id"] = user.id
            session["username"] = user.username
            session["role"] = user.role
            session["full_name"] = user.full_name

            # 记录登录日志
            log = OperationLog(
                user_id=user.id,
                action="login",
                module="system"
            )
            db.session.add(log)
            db.session.commit()

            flash(f"欢迎 {user.full_name}", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("用户名或密码错误", "danger")

    return render_template("auth/login.html")


def _find_user_by_identifier(identifier):
    """通过用户名/学号/教工号查找用户。"""
    if not identifier:
        return None

    user = User.query.filter_by(username=identifier).first()
    if user:
        return user

    student = Student.query.filter_by(student_no=identifier).first()
    if student and student.user_id:
        return student.user

    teacher = Teacher.query.filter_by(teacher_no=identifier).first()
    if teacher and teacher.user_id:
        return teacher.user

    return None


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        identifier = request.form.get("identifier", "").strip()

        if not identifier:
            flash("请输入学号、工号或用户名", "danger")
            return render_template("auth/forgot_password.html", identifier=identifier)

        user = _find_user_by_identifier(identifier)
        if not user:
            flash("未找到对应账号，请检查后重试", "danger")
            return render_template("auth/forgot_password.html", identifier=identifier)

        key = secrets.token_hex(3).upper()
        expires_at = datetime.utcnow() + timedelta(minutes=10)

        try:
            db.session.add(PasswordResetKey(
                user_id=user.id,
                reset_key=key,
                requested_identifier=identifier,
                expires_at=expires_at,
                request_ip=request.remote_addr,
            ))

            # 通过管理员操作日志回传key，管理员可在“操作日志”页面查看。
            db.session.add(OperationLog(
                user_id=None,
                action="password_reset_key",
                module="auth",
                details=(
                    f"忘记密码申请：账号 {user.username}（标识 {identifier}）"
                    f"，重置Key：{key}，有效期至 {expires_at.strftime('%Y-%m-%d %H:%M:%S')}"
                )
            ))
            db.session.commit()
            flash("申请已提交，请联系系统管理员获取重置Key（10分钟内有效）", "success")
            return redirect(url_for("reset_password_with_key"))
        except Exception:
            db.session.rollback()
            flash("申请失败，请稍后重试", "danger")

    return render_template("auth/forgot_password.html", identifier="")


@app.route("/reset-password-with-key", methods=["GET", "POST"])
def reset_password_with_key():
    if request.method == "POST":
        account = request.form.get("account", "").strip()
        key = request.form.get("key", "").strip().upper()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not account or not key or not new_password or not confirm_password:
            flash("请完整填写账号、Key 和新密码", "danger")
            return render_template("auth/reset_password_with_key.html", account=account, key=key)

        if new_password != confirm_password:
            flash("两次输入密码不一致", "danger")
            return render_template("auth/reset_password_with_key.html", account=account, key=key)

        if len(new_password) < 6:
            flash("密码至少 6 个字符", "danger")
            return render_template("auth/reset_password_with_key.html", account=account, key=key)

        user = _find_user_by_identifier(account)
        if not user:
            flash("账号不存在", "danger")
            return render_template("auth/reset_password_with_key.html", account=account, key=key)

        reset_record = (
            PasswordResetKey.query
            .filter_by(user_id=user.id, reset_key=key, is_used=False)
            .order_by(PasswordResetKey.created_at.desc())
            .first()
        )

        now = datetime.utcnow()
        if not reset_record:
            flash("Key 无效，请重新申请", "danger")
            return render_template("auth/reset_password_with_key.html", account=account, key=key)

        if reset_record.expires_at < now:
            flash("Key 已过期，请重新申请", "danger")
            return render_template("auth/reset_password_with_key.html", account=account, key="")

        try:
            user.set_password(new_password)
            reset_record.is_used = True
            reset_record.used_at = now

            db.session.add(OperationLog(
                user_id=user.id,
                action="reset_password_with_key",
                module="auth",
                details=f"账号 {user.username} 通过重置Key修改密码"
            ))
            db.session.commit()
            flash("密码重置成功，请使用新密码登录", "success")
            return redirect(url_for("login"))
        except Exception:
            db.session.rollback()
            flash("密码重置失败，请稍后重试", "danger")

    return render_template("auth/reset_password_with_key.html", account="", key="")


@app.route("/logout")
def logout():
    session.clear()
    flash("已退出登录", "info")
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        old_password = request.form.get("old_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        user = User.query.get(session["user_id"])
        if not user.check_password(old_password):
            flash("原密码错误", "danger")
        elif new_password != confirm_password:
            flash("两次输入密码不一致", "danger")
        elif len(new_password) < 6:
            flash("密码至少 6 个字符", "danger")
        else:
            user.set_password(new_password)
            db.session.commit()
            flash("密码修改成功", "success")
            return redirect(url_for("dashboard"))

    return render_template("auth/change_password.html")


# ======================== 首页 / 仪表板 ========================

@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    user = User.query.get(session["user_id"])
    
    stats = {
        "total_students": Student.query.count(),
        "total_teachers": Teacher.query.count(),
        "total_courses": Course.query.count(),
        "total_classes": Class.query.count(),
    }

    return render_template("dashboard.html", user=user, stats=stats)


# ======================== 学生管理路由 ========================

def _teacher_managed_class_ids(user_id):
    """返回教师（按用户ID）负责的班级ID列表。"""
    rows = Class.query.filter_by(headteacher_id=user_id).with_entities(Class.id).all()
    return [r[0] for r in rows]


def _teacher_course_ids(user_id):
    """返回教师（按用户ID）可授课课程ID列表（主讲+协同）。"""
    teacher = Teacher.query.filter_by(user_id=user_id).first()
    if not teacher:
        return []

    primary_rows = Course.query.filter_by(teacher_id=teacher.id).with_entities(Course.id).all()
    secondary_rows = (
        db.session.query(Course.id)
        .join(course_teacher_links, Course.id == course_teacher_links.c.course_id)
        .filter(course_teacher_links.c.teacher_id == teacher.id)
        .all()
    )
    return list({r[0] for r in primary_rows + secondary_rows})


def _is_counselor(user_id):
    """是否为辅导员（即至少负责一个班级）。"""
    if not user_id:
        return False
    return Class.query.filter_by(headteacher_id=user_id).first() is not None


@app.context_processor
def inject_role_flags():
    role = session.get("role")
    user_id = session.get("user_id")
    return {
        "is_counselor": role == "teacher" and _is_counselor(user_id)
    }

@app.route("/students")
@login_required
def list_students():
    if session.get("role") == "teacher" and not _is_counselor(session.get("user_id")):
        flash("任课老师只能管理成绩，不能进行班级管理", "warning")
        return redirect(url_for("list_scores"))

    query = request.args.get("q", "").strip()
    class_id = request.args.get("class_id", type=int)
    status = request.args.get("status", "").strip()

    q = Student.query
    classes_query = Class.query

    if session.get("role") == "teacher":
        managed_ids = _teacher_managed_class_ids(session["user_id"])
        if managed_ids:
            q = q.filter(Student.class_id.in_(managed_ids))
            classes_query = classes_query.filter(Class.id.in_(managed_ids))
        else:
            q = q.filter(Student.id == -1)
            classes_query = classes_query.filter(Class.id == -1)

    if query:
        like_query = f"%{query}%"
        q = q.filter(
            db.or_(
                Student.student_no.ilike(like_query),
                Student.name.ilike(like_query),
                Student.major.ilike(like_query)
            )
        )

    if class_id:
        if session.get("role") == "teacher":
            managed_ids = _teacher_managed_class_ids(session["user_id"])
            if class_id not in managed_ids:
                flash("你只能查看自己负责班级的学生", "error")
                return redirect(url_for("list_students"))
        q = q.filter(Student.class_id == class_id)

    if status:
        q = q.filter(Student.status == status)

    students = q.order_by(Student.id.desc()).all()
    classes = classes_query.all()

    return render_template(
        "students/list.html",
        students=students,
        classes=classes,
        query=query,
        class_id=class_id,
        status=status
    )


@app.route("/students/add", methods=["GET", "POST"])
@role_required("admin", "teacher")
def add_student():
    classes = Class.query.all()
    managed_ids = []
    if session.get("role") == "teacher":
        managed_ids = _teacher_managed_class_ids(session["user_id"])
        classes = Class.query.filter(Class.id.in_(managed_ids)).all() if managed_ids else []

    if request.method == "POST":
        student_no = request.form.get("student_no", "").strip()
        name = request.form.get("name", "").strip()
        gender = request.form.get("gender", "").strip()
        age_raw = request.form.get("age", "").strip()
        class_id = request.form.get("class_id", type=int)
        status = request.form.get("status", "enrolled").strip() or "enrolled"
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip()
        major = request.form.get("major", "").strip()
        id_number = request.form.get("id_number", "").strip()

        errors = []
        if not student_no:
            errors.append("学号不能为空")
        if not name:
            errors.append("姓名不能为空")
        if age_raw and not age_raw.isdigit():
            errors.append("年龄必须是数字")
        if status not in ["enrolled", "leave", "dropout"]:
            errors.append("学生状态不合法")

        if errors:
            for error in errors:
                flash(error, "danger")
            return render_template("students/form.html", student=Student(), classes=classes, action="add")

        if Student.query.filter_by(student_no=student_no).first():
            flash("学号已存在", "error")
            student = Student(
                student_no=student_no,
                name=name,
                gender=gender,
                age=int(age_raw) if age_raw and age_raw.isdigit() else None,
                class_id=class_id,
                status=status,
                phone=phone,
                email=email,
                major=major,
                id_number=id_number,
            )
            return render_template("students/form.html", student=student, classes=classes, action="add")

        if User.query.filter_by(username=student_no).first():
            flash("该学号已被占用为系统用户名，请使用其他学号", "error")
            student = Student(
                student_no=student_no,
                name=name,
                gender=gender,
                age=int(age_raw) if age_raw and age_raw.isdigit() else None,
                class_id=class_id,
                phone=phone,
                email=email,
                major=major,
                id_number=id_number,
            )
            return render_template("students/form.html", student=student, classes=classes, action="add")

        if session.get("role") == "teacher":
            if not class_id or class_id not in managed_ids:
                flash("教师只能将学生添加到自己负责的班级", "error")
                student = Student(
                    student_no=student_no,
                    name=name,
                    gender=gender,
                    age=int(age_raw) if age_raw and age_raw.isdigit() else None,
                    class_id=class_id,
                    status=status,
                    phone=phone,
                    email=email,
                    major=major,
                    id_number=id_number,
                )
                return render_template("students/form.html", student=student, classes=classes, action="add")

        try:
            user = User(
                username=student_no,
                full_name=name,
                role="student",
                email=email or None,
                phone=phone or None,
                is_active=True,
            )
            user.set_password("123456")
            db.session.add(user)
            db.session.flush()

            student = Student(
                user_id=user.id,
                student_no=student_no,
                name=name,
                gender=gender,
                age=int(age_raw) if age_raw else None,
                class_id=class_id,
                status=status,
                phone=phone,
                email=email,
                major=major,
                id_number=id_number,
            )
            db.session.add(student)
            db.session.commit()

            # 记录操作日志
            log = OperationLog(
                user_id=session["user_id"],
                action="create",
                module="student",
                record_id=student.id,
                details=f"新增学生 {name}({student_no}) 并创建学生账号"
            )
            db.session.add(log)
            db.session.commit()

            flash("新增学生成功，已同步创建学生账号（用户名：学号，初始密码：123456）", "success")
            return redirect(url_for("list_students"))
        except Exception as e:
            db.session.rollback()
            flash(f"添加失败：{str(e)}", "danger")

    return render_template("students/form.html", student=Student(), classes=classes, action="add")


@app.route("/students/edit/<int:student_id>", methods=["GET", "POST"])
@role_required("admin", "teacher")
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    classes = Class.query.all()
    managed_ids = []

    if session.get("role") == "teacher":
        managed_ids = _teacher_managed_class_ids(session["user_id"])
        if student.class_id not in managed_ids:
            flash("你只能编辑自己负责班级的学生", "error")
            return redirect(url_for("list_students"))
        classes = Class.query.filter(Class.id.in_(managed_ids)).all() if managed_ids else []

    if request.method == "POST":
        student_no = request.form.get("student_no", "").strip()
        name = request.form.get("name", "").strip()
        gender = request.form.get("gender", "").strip()
        age_raw = request.form.get("age", "").strip()
        class_id = request.form.get("class_id", type=int)
        status = request.form.get("status", "enrolled").strip() or "enrolled"
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip()
        major = request.form.get("major", "").strip()
        id_number = request.form.get("id_number", "").strip()

        if not student_no or not name:
            flash("学号和姓名不能为空", "error")
            student.student_no = student_no
            student.name = name
            student.gender = gender
            student.age = int(age_raw) if age_raw.isdigit() else None
            student.class_id = class_id
            student.status = status
            student.phone = phone
            student.email = email
            student.major = major
            student.id_number = id_number
            return render_template("students/form.html", student=student, classes=classes, action="edit")

        if age_raw and not age_raw.isdigit():
            flash("年龄必须是数字", "error")
            student.student_no = student_no
            student.name = name
            student.gender = gender
            student.age = None
            student.class_id = class_id
            student.status = status
            student.phone = phone
            student.email = email
            student.major = major
            student.id_number = id_number
            return render_template("students/form.html", student=student, classes=classes, action="edit")

        if status not in ["enrolled", "leave", "dropout"]:
            flash("学生状态不合法", "error")
            student.student_no = student_no
            student.name = name
            student.gender = gender
            student.age = int(age_raw) if age_raw else None
            student.class_id = class_id
            student.status = status
            student.phone = phone
            student.email = email
            student.major = major
            student.id_number = id_number
            return render_template("students/form.html", student=student, classes=classes, action="edit")

        student_no_exists = Student.query.filter(
            Student.student_no == student_no,
            Student.id != student.id
        ).first()
        if student_no_exists:
            flash("学号已存在，请使用其他学号", "error")
            student.student_no = student_no
            student.name = name
            student.gender = gender
            student.age = int(age_raw) if age_raw else None
            student.class_id = class_id
            student.status = status
            student.phone = phone
            student.email = email
            student.major = major
            student.id_number = id_number
            return render_template("students/form.html", student=student, classes=classes, action="edit")

        linked_user = student.user
        username_holder = User.query.filter_by(username=student_no).first()
        if username_holder and (not linked_user or username_holder.id != linked_user.id):
            flash("该学号已被占用为系统用户名，请使用其他学号", "error")
            student.student_no = student_no
            student.name = name
            student.gender = gender
            student.age = int(age_raw) if age_raw else None
            student.class_id = class_id
            student.status = status
            student.phone = phone
            student.email = email
            student.major = major
            student.id_number = id_number
            return render_template("students/form.html", student=student, classes=classes, action="edit")

        student.student_no = student_no
        student.name = name
        student.gender = gender
        student.age = int(age_raw) if age_raw else None
        student.class_id = class_id
        student.status = status
        student.phone = phone
        student.email = email
        student.major = major
        student.id_number = id_number

        if session.get("role") == "teacher" and student.class_id not in managed_ids:
            flash("教师不能把学生调整到非本人负责的班级", "error")
            return render_template("students/form.html", student=student, classes=classes, action="edit")

        try:
            if not linked_user:
                linked_user = User(
                    username=student.student_no,
                    full_name=student.name,
                    role="student",
                    email=student.email or None,
                    phone=student.phone or None,
                    is_active=True,
                )
                linked_user.set_password("123456")
                db.session.add(linked_user)
                db.session.flush()
                student.user_id = linked_user.id
            else:
                linked_user.username = student.student_no
                linked_user.full_name = student.name
                linked_user.email = student.email or None
                linked_user.phone = student.phone or None
                linked_user.role = "student"

            db.session.commit()

            log = OperationLog(
                user_id=session["user_id"],
                action="update",
                module="student",
                record_id=student.id,
                details=f"修改学生信息 {student.name}({student.student_no})"
            )
            db.session.add(log)
            db.session.commit()

            flash("修改成功，学生账号信息已同步", "success")
            return redirect(url_for("list_students"))
        except Exception as e:
            db.session.rollback()
            flash(f"修改失败：{str(e)}", "danger")

    return render_template("students/form.html", student=student, classes=classes, action="edit")


@app.route("/students/delete/<int:student_id>", methods=["POST"])
@role_required("admin")
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    name = student.name
    linked_user = student.user

    try:
        log = OperationLog(
            user_id=session["user_id"],
            action="delete",
            module="student",
            record_id=student.id,
            details=f"删除学生 {name}({student.student_no})"
        )
        db.session.add(log)
        db.session.delete(student)
        if linked_user:
            db.session.delete(linked_user)
        db.session.commit()

        flash("删除成功", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"删除失败：{str(e)}", "danger")

    return redirect(url_for("list_students"))


# ======================== 班级管理路由 ========================

@app.route("/classes")
@login_required
def list_classes():
    q = Class.query
    if session.get("role") == "teacher":
        if not _is_counselor(session["user_id"]):
            flash("任课老师不能进行班级管理", "warning")
            return redirect(url_for("list_scores"))

        managed_ids = _teacher_managed_class_ids(session["user_id"])
        if managed_ids:
            q = q.filter(Class.id.in_(managed_ids))
        else:
            q = q.filter(Class.id == -1)

    classes = q.order_by(Class.id.desc()).all()
    teachers = Teacher.query.all()
    return render_template("classes/list.html", classes=classes, teachers=teachers)


@app.route("/classes/add", methods=["GET", "POST"])
@role_required("admin")
def add_class():
    teachers = (
        Teacher.query
        .join(User, Teacher.user_id == User.id)
        .filter(User.role == "teacher")
        .all()
    )

    def build_class_from_form():
        return Class(
            class_no=request.form.get("class_no", "").strip(),
            class_name=request.form.get("class_name", "").strip(),
            headteacher_id=request.form.get("headteacher_id", type=int),
            grade=request.form.get("grade", "").strip(),
        )

    if request.method == "POST":
        selected_teacher_id = request.form.get("headteacher_id", type=int)
        cls = build_class_from_form()

        if not cls.class_no or not cls.class_name:
            flash("班级编号和班级名称不能为空", "error")
            return render_template("classes/form.html", teachers=teachers, cls=cls, action="add")

        if len(cls.class_no) > 50:
            flash("班级编号长度不能超过 50 个字符", "error")
            return render_template("classes/form.html", teachers=teachers, cls=cls, action="add")

        if len(cls.class_name) > 100:
            flash("班级名称长度不能超过 100 个字符", "error")
            return render_template("classes/form.html", teachers=teachers, cls=cls, action="add")

        if Class.query.filter_by(class_no=cls.class_no).first():
            flash("班级编号已存在，请使用其他编号", "error")
            return render_template(
                "classes/form.html",
                teachers=teachers,
                cls=cls,
                selected_teacher_id=selected_teacher_id,
                action="add",
            )

        # 表单提交的是 teachers.id；Class.headteacher_id 需要存 users.id。
        if selected_teacher_id:
            teacher_obj = Teacher.query.get(selected_teacher_id)
            if not teacher_obj:
                flash("所选辅导员不存在，请重新选择", "error")
                return render_template(
                    "classes/form.html",
                    teachers=teachers,
                    cls=cls,
                    selected_teacher_id=selected_teacher_id,
                    action="add",
                )
            if not teacher_obj.user_id:
                flash("所选教师未绑定系统账号，暂时不能设为辅导员", "error")
                return render_template(
                    "classes/form.html",
                    teachers=teachers,
                    cls=cls,
                    selected_teacher_id=selected_teacher_id,
                    action="add",
                )
            if not teacher_obj.user or teacher_obj.user.role != "teacher":
                flash("辅导员必须是教师角色账号", "error")
                return render_template(
                    "classes/form.html",
                    teachers=teachers,
                    cls=cls,
                    selected_teacher_id=selected_teacher_id,
                    action="add",
                )
            cls.headteacher_id = teacher_obj.user_id
        else:
            cls.headteacher_id = None

        try:
            db.session.add(cls)
            db.session.commit()

            log = OperationLog(
                user_id=session["user_id"],
                action="create",
                module="class",
                record_id=cls.id,
                details=f"新增班级 {cls.class_name}"
            )
            db.session.add(log)
            db.session.commit()

            flash("新增班级成功", "success")
            return redirect(url_for("list_classes"))
        except Exception:
            db.session.rollback()
            flash("新增班级失败，请检查输入内容后重试", "error")
            return render_template(
                "classes/form.html",
                teachers=teachers,
                cls=cls,
                selected_teacher_id=selected_teacher_id,
                action="add",
            )

    return render_template(
        "classes/form.html",
        teachers=teachers,
        cls=Class(),
        selected_teacher_id=None,
        action="add",
    )


@app.route("/classes/edit/<int:class_id>", methods=["GET", "POST"])
@role_required("admin")
def edit_class(class_id):
    teachers = (
        Teacher.query
        .join(User, Teacher.user_id == User.id)
        .filter(User.role == "teacher")
        .all()
    )
    cls = Class.query.get_or_404(class_id)

    selected_teacher_id = None
    if cls.headteacher_id:
        bound_teacher = Teacher.query.filter_by(user_id=cls.headteacher_id).first()
        if bound_teacher:
            selected_teacher_id = bound_teacher.id

    if request.method == "POST":
        selected_teacher_id = request.form.get("headteacher_id", type=int)
        class_no = request.form.get("class_no", "").strip()
        class_name = request.form.get("class_name", "").strip()
        grade = request.form.get("grade", "").strip()
        headteacher_id = None

        if selected_teacher_id:
            teacher_obj = Teacher.query.get(selected_teacher_id)
            if not teacher_obj:
                flash("所选辅导员不存在，请重新选择", "error")
                return render_template(
                    "classes/form.html",
                    teachers=teachers,
                    cls=cls,
                    selected_teacher_id=selected_teacher_id,
                    action="edit",
                )
            if not teacher_obj.user_id:
                flash("所选教师未绑定系统账号，暂时不能设为辅导员", "error")
                return render_template(
                    "classes/form.html",
                    teachers=teachers,
                    cls=cls,
                    selected_teacher_id=selected_teacher_id,
                    action="edit",
                )
            if not teacher_obj.user or teacher_obj.user.role != "teacher":
                flash("辅导员必须是教师角色账号", "error")
                return render_template(
                    "classes/form.html",
                    teachers=teachers,
                    cls=cls,
                    selected_teacher_id=selected_teacher_id,
                    action="edit",
                )
            headteacher_id = teacher_obj.user_id

        cls.class_no = class_no
        cls.class_name = class_name
        cls.grade = grade
        cls.headteacher_id = headteacher_id

        if not cls.class_no or not cls.class_name:
            flash("班级编号和班级名称不能为空", "error")
            return render_template(
                "classes/form.html",
                teachers=teachers,
                cls=cls,
                selected_teacher_id=selected_teacher_id,
                action="edit",
            )

        if len(cls.class_no) > 50:
            flash("班级编号长度不能超过 50 个字符", "error")
            return render_template(
                "classes/form.html",
                teachers=teachers,
                cls=cls,
                selected_teacher_id=selected_teacher_id,
                action="edit",
            )

        if len(cls.class_name) > 100:
            flash("班级名称长度不能超过 100 个字符", "error")
            return render_template(
                "classes/form.html",
                teachers=teachers,
                cls=cls,
                selected_teacher_id=selected_teacher_id,
                action="edit",
            )

        exists = Class.query.filter(
            Class.class_no == cls.class_no,
            Class.id != cls.id,
        ).first()
        if exists:
            flash("班级编号已存在，请使用其他编号", "error")
            return render_template(
                "classes/form.html",
                teachers=teachers,
                cls=cls,
                selected_teacher_id=selected_teacher_id,
                action="edit",
            )

        try:
            db.session.commit()

            log = OperationLog(
                user_id=session["user_id"],
                action="update",
                module="class",
                record_id=cls.id,
                details=f"修改班级 {cls.class_name}({cls.class_no})",
            )
            db.session.add(log)
            db.session.commit()

            flash("班级修改成功", "success")
            return redirect(url_for("list_classes"))
        except Exception:
            db.session.rollback()
            flash("班级修改失败，请检查输入内容后重试", "error")
            return render_template(
                "classes/form.html",
                teachers=teachers,
                cls=cls,
                selected_teacher_id=selected_teacher_id,
                action="edit",
            )

    return render_template(
        "classes/form.html",
        teachers=teachers,
        cls=cls,
        selected_teacher_id=selected_teacher_id,
        action="edit",
    )


@app.route("/classes/delete/<int:class_id>", methods=["POST"])
@role_required("admin")
def delete_class(class_id):
    cls = Class.query.get_or_404(class_id)

    try:
        student_count = Student.query.filter_by(class_id=cls.id).count()
        if student_count > 0:
            flash("该班级下还有学生，无法删除，请先转移或删除学生", "error")
            return redirect(url_for("list_classes"))

        log = OperationLog(
            user_id=session["user_id"],
            action="delete",
            module="class",
            record_id=cls.id,
            details=f"删除班级 {cls.class_name}({cls.class_no})"
        )
        db.session.add(log)
        db.session.delete(cls)
        db.session.commit()

        flash("删除成功", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"删除失败：{str(e)}", "danger")

    return redirect(url_for("list_classes"))


# ======================== 教师管理路由 ========================

@app.route("/teachers")
@login_required
def list_teachers():
    query = request.args.get("q", "").strip()
    q = Teacher.query

    if query:
        like_query = f"%{query}%"
        q = q.filter(
            db.or_(
                Teacher.teacher_no.ilike(like_query),
                Teacher.name.ilike(like_query),
                Teacher.department.ilike(like_query)
            )
        )

    teachers = q.order_by(Teacher.id.desc()).all()
    return render_template("teachers/list.html", teachers=teachers, query=query)


@app.route("/teachers/add", methods=["GET", "POST"])
@role_required("admin")
def add_teacher():
    def build_teacher_from_form():
        return Teacher(
            teacher_no=request.form.get("teacher_no", "").strip(),
            name=request.form.get("name", "").strip(),
            gender=request.form.get("gender", "").strip(),
            department=request.form.get("department", "").strip(),
            phone=request.form.get("phone", "").strip(),
            email=request.form.get("email", "").strip(),
            qualification=request.form.get("qualification", "").strip(),
        )

    if request.method == "POST":
        teacher = build_teacher_from_form()

        if not teacher.teacher_no or not teacher.name:
            flash("教工号和姓名不能为空", "error")
            return render_template("teachers/form.html", teacher=teacher, action="add")

        if len(teacher.teacher_no) > 50:
            flash("教工号长度不能超过 50 个字符", "error")
            return render_template("teachers/form.html", teacher=teacher, action="add")

        if len(teacher.name) > 100:
            flash("姓名长度不能超过 100 个字符", "error")
            return render_template("teachers/form.html", teacher=teacher, action="add")

        if teacher.phone and len(teacher.phone) > 20:
            flash("电话号码长度不能超过 20 位", "error")
            return render_template("teachers/form.html", teacher=teacher, action="add")

        if teacher.email and len(teacher.email) > 100:
            flash("邮箱长度不能超过 100 个字符", "error")
            return render_template("teachers/form.html", teacher=teacher, action="add")

        if Teacher.query.filter_by(teacher_no=teacher.teacher_no).first():
            flash("教工号已存在，请使用其他教工号", "error")
            return render_template("teachers/form.html", teacher=teacher, action="add")

        if User.query.filter_by(username=teacher.teacher_no).first():
            flash("该教工号已被占用为系统用户名，请使用其他教工号", "error")
            return render_template("teachers/form.html", teacher=teacher, action="add")

        try:
            user = User(
                username=teacher.teacher_no,
                full_name=teacher.name,
                role="teacher",
                email=teacher.email or None,
                phone=teacher.phone or None,
                is_active=True,
            )
            user.set_password("123456")
            db.session.add(user)
            db.session.flush()

            teacher.user_id = user.id
            db.session.add(teacher)
            db.session.commit()

            log = OperationLog(
                user_id=session["user_id"],
                action="create",
                module="teacher",
                record_id=teacher.id,
                details=f"新增教师 {teacher.name}({teacher.teacher_no}) 并创建教师账号"
            )
            db.session.add(log)
            db.session.commit()

            flash("新增教师成功，已同步创建教师账号（用户名：教工号，初始密码：123456）", "success")
            return redirect(url_for("list_teachers"))
        except Exception:
            db.session.rollback()
            flash("新增教师失败，请检查输入内容后重试", "error")
            return render_template("teachers/form.html", teacher=teacher, action="add")

    return render_template("teachers/form.html", teacher=Teacher(), action="add")


@app.route("/teachers/edit/<int:teacher_id>", methods=["GET", "POST"])
@role_required("admin")
def edit_teacher(teacher_id):
    teacher = Teacher.query.get_or_404(teacher_id)

    if request.method == "POST":
        teacher_no = request.form.get("teacher_no", "").strip()
        name = request.form.get("name", "").strip()
        gender = request.form.get("gender", "").strip()
        department = request.form.get("department", "").strip()
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip()
        qualification = request.form.get("qualification", "").strip()

        if not teacher_no or not name:
            flash("教工号和姓名不能为空", "error")
            teacher.teacher_no = teacher_no
            teacher.name = name
            teacher.gender = gender
            teacher.department = department
            teacher.phone = phone
            teacher.email = email
            teacher.qualification = qualification
            return render_template("teachers/form.html", teacher=teacher, action="edit")

        teacher_no_exists = Teacher.query.filter(
            Teacher.teacher_no == teacher_no,
            Teacher.id != teacher.id
        ).first()
        if teacher_no_exists:
            flash("教工号已存在，请使用其他教工号", "error")
            teacher.teacher_no = teacher_no
            teacher.name = name
            teacher.gender = gender
            teacher.department = department
            teacher.phone = phone
            teacher.email = email
            teacher.qualification = qualification
            return render_template("teachers/form.html", teacher=teacher, action="edit")

        linked_user = teacher.user
        username_holder = User.query.filter_by(username=teacher_no).first()
        if username_holder and (not linked_user or username_holder.id != linked_user.id):
            flash("该教工号已被占用为系统用户名，请使用其他教工号", "error")
            teacher.teacher_no = teacher_no
            teacher.name = name
            teacher.gender = gender
            teacher.department = department
            teacher.phone = phone
            teacher.email = email
            teacher.qualification = qualification
            return render_template("teachers/form.html", teacher=teacher, action="edit")

        teacher.teacher_no = teacher_no
        teacher.name = name
        teacher.gender = gender
        teacher.department = department
        teacher.phone = phone
        teacher.email = email
        teacher.qualification = qualification

        try:
            if not linked_user:
                linked_user = User(
                    username=teacher.teacher_no,
                    full_name=teacher.name,
                    role="teacher",
                    email=teacher.email or None,
                    phone=teacher.phone or None,
                    is_active=True,
                )
                linked_user.set_password("123456")
                db.session.add(linked_user)
                db.session.flush()
                teacher.user_id = linked_user.id
            else:
                linked_user.username = teacher.teacher_no
                linked_user.full_name = teacher.name
                linked_user.email = teacher.email or None
                linked_user.phone = teacher.phone or None
                linked_user.role = "teacher"

            db.session.commit()

            log = OperationLog(
                user_id=session["user_id"],
                action="update",
                module="teacher",
                record_id=teacher.id,
                details=f"修改教师信息 {teacher.name}({teacher.teacher_no})"
            )
            db.session.add(log)
            db.session.commit()

            flash("修改成功，教师账号信息已同步", "success")
            return redirect(url_for("list_teachers"))
        except Exception as e:
            db.session.rollback()
            flash(f"修改失败：{str(e)}", "danger")

    return render_template("teachers/form.html", teacher=teacher, action="edit")


@app.route("/teachers/delete/<int:teacher_id>", methods=["POST"])
@role_required("admin")
def delete_teacher(teacher_id):
    teacher = Teacher.query.get_or_404(teacher_id)
    name = teacher.name
    linked_user = teacher.user

    if linked_user and linked_user.id == session.get("user_id"):
        flash("不能删除当前登录账号对应的教师信息", "error")
        return redirect(url_for("list_teachers"))

    try:
        log = OperationLog(
            user_id=session["user_id"],
            action="delete",
            module="teacher",
            record_id=teacher.id,
            details=f"删除教师 {name}({teacher.teacher_no})"
        )
        db.session.add(log)

        # 先清理班级上的辅导员引用，避免后续删除教师账号时产生关联问题。
        Class.query.filter_by(headteacher_id=teacher.user_id).update({"headteacher_id": None})

        user_deleted = False
        if linked_user:
            has_related_courses = Course.query.filter_by(teacher_id=teacher.id).count() > 0
            if has_related_courses:
                linked_user.is_active = False
                linked_user.role = "teacher"
            else:
                db.session.delete(linked_user)
                user_deleted = True

        db.session.delete(teacher)
        db.session.commit()

        if linked_user and not user_deleted:
            flash("删除成功，教师账号已自动禁用", "success")
        else:
            flash("删除成功", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"删除失败：{str(e)}", "danger")

    return redirect(url_for("list_teachers"))


# ======================== 课程管理路由 ========================

@app.route("/courses")
@login_required
def list_courses():
    semester = request.args.get("semester", "").strip()
    status = request.args.get("status", "").strip()

    q = Course.query

    if session.get("role") == "teacher":
        teacher_course_ids = _teacher_course_ids(session["user_id"])
        if teacher_course_ids:
            q = q.filter(Course.id.in_(teacher_course_ids))
        else:
            q = q.filter(Course.id == -1)

    if semester:
        q = q.filter(Course.semester == semester)

    if status:
        q = q.filter(Course.status == status)

    courses = q.order_by(Course.id.desc()).all()
    semesters = db.session.query(Course.semester).distinct().all()

    return render_template(
        "courses/list.html",
        courses=courses,
        semesters=[s[0] for s in semesters if s[0]],
        semester=semester,
        status=status
    )


@app.route("/courses/add", methods=["GET", "POST"])
@role_required("admin")
def add_course():
    teachers = Teacher.query.all()

    if request.method == "POST":
        course_no = request.form.get("course_no", "").strip()
        course_name = request.form.get("course_name", "").strip()
        credits = request.form.get("credits", type=float)
        hours = request.form.get("hours", type=int)
        teacher_id = request.form.get("teacher_id", type=int)
        co_teacher_ids = request.form.getlist("co_teacher_ids")
        max_capacity = request.form.get("max_capacity", type=int, default=50)
        semester = request.form.get("semester", "").strip()

        if not course_no or not course_name:
            flash("课程编号和课程名称不能为空", "danger")
            return render_template("courses/form.html", teachers=teachers, action="add")

        try:
            course = Course(
                course_no=course_no,
                course_name=course_name,
                credits=credits,
                hours=hours,
                teacher_id=teacher_id,
                max_capacity=max_capacity,
                semester=semester,
                status="open"
            )
            db.session.add(course)
            db.session.flush()

            co_ids = []
            for raw in co_teacher_ids:
                try:
                    tid = int(raw)
                except (TypeError, ValueError):
                    continue
                if tid and tid != teacher_id:
                    co_ids.append(tid)
            if co_ids:
                course.co_teachers = Teacher.query.filter(Teacher.id.in_(list(set(co_ids)))).all()

            db.session.commit()

            log = OperationLog(
                user_id=session["user_id"],
                action="create",
                module="course",
                record_id=course.id,
                details=f"新增课程 {course_name}"
            )
            db.session.add(log)
            db.session.commit()

            flash("新增课程成功", "success")
            return redirect(url_for("list_courses"))
        except Exception as e:
            db.session.rollback()
            flash(f"添加失败：{str(e)}", "danger")

    return render_template("courses/form.html", teachers=teachers, action="add")


# ======================== 选课管理 ========================

@app.route("/enrollments")
@login_required
def list_enrollments():
    if session.get("role") == "teacher" and not _is_counselor(session.get("user_id")):
        flash("任课老师只能管理成绩，不能进行班级管理", "warning")
        return redirect(url_for("list_scores"))

    status = request.args.get("status", "").strip()
    q = Enrollment.query.join(Course, Enrollment.course_id == Course.id).join(Student, Enrollment.student_id == Student.id)

    if session.get("role") == "student":
        student = Student.query.filter_by(user_id=session["user_id"]).first()
        if not student:
            flash("当前账号未绑定学生档案，请联系管理员处理", "warning")
            return redirect(url_for("dashboard"))
        q = q.filter(Enrollment.student_id == student.id)
    elif session.get("role") == "teacher":
        managed_ids = _teacher_managed_class_ids(session["user_id"])
        if managed_ids:
            q = q.filter(Student.class_id.in_(managed_ids))
        else:
            q = q.filter(Enrollment.id == -1)
    else:
        q = q

    if status:
        q = q.filter(Enrollment.status == status)

    enrollments = q.order_by(Enrollment.enrollment_date.desc()).all()

    return render_template("enrollments/list.html", enrollments=enrollments, status=status)


def _submit_enrollment(student, course, operator_user_id, actor_label):
    existing = Enrollment.query.filter_by(student_id=student.id, course_id=course.id).first()

    if existing and existing.status == "enrolled":
        return False, "该学生已经选过这门课程", "warning"

    current_count = db.session.query(Enrollment).filter_by(
        course_id=course.id,
        status="enrolled",
    ).count()
    if current_count >= course.max_capacity:
        return False, "课程已满员", "danger"

    try:
        if existing and existing.status == "withdrawn":
            enrollment = existing
            enrollment.status = "enrolled"
            enrollment.completion_date = None
            enrollment.enrollment_date = datetime.utcnow()
            log_action = "re_enroll"
            log_details = f"{actor_label}：重新选课 {student.name}-{course.course_name}"
        else:
            enrollment = Enrollment(
                student_id=student.id,
                course_id=course.id,
                status="enrolled"
            )
            db.session.add(enrollment)
            log_action = "enroll"
            log_details = f"{actor_label}：选课 {student.name}-{course.course_name}"

        db.session.commit()

        db.session.add(OperationLog(
            user_id=operator_user_id,
            action=log_action,
            module="enrollment",
            record_id=enrollment.id,
            details=log_details,
        ))
        db.session.commit()
        return True, "选课成功", "success"
    except Exception as e:
        db.session.rollback()
        return False, f"选课失败：{str(e)}", "danger"


@app.route("/enrollments/add/<int:course_id>", methods=["POST"])
@login_required
def add_enrollment(course_id):
    if session.get("role") != "student":
        flash("只有学生账号可以直接选课", "warning")
        return redirect(url_for("list_courses"))

    course = Course.query.get_or_404(course_id)
    student = Student.query.filter_by(user_id=session["user_id"]).first()

    if not student:
        flash("学生信息不存在", "danger")
        return redirect(url_for("list_courses"))

    success, message, category = _submit_enrollment(
        student=student,
        course=course,
        operator_user_id=session["user_id"],
        actor_label="学生自助选课",
    )
    flash(message, category)

    return redirect(url_for("list_courses"))


@app.route("/enrollments/admin-add", methods=["GET", "POST"])
@role_required("admin")
def admin_add_enrollment():
    students = Student.query.order_by(Student.student_no.asc()).all()
    courses = Course.query.order_by(Course.id.desc()).all()

    selected_student_id = request.form.get("student_id", type=int) if request.method == "POST" else None
    selected_course_id = request.form.get("course_id", type=int) if request.method == "POST" else None

    if request.method == "POST":
        if not selected_student_id or not selected_course_id:
            flash("请选择学生和课程", "danger")
            return render_template(
                "enrollments/admin_form.html",
                students=students,
                courses=courses,
                selected_student_id=selected_student_id,
                selected_course_id=selected_course_id,
            )

        student = Student.query.get(selected_student_id)
        course = Course.query.get(selected_course_id)
        if not student or not course:
            flash("学生或课程不存在", "danger")
            return render_template(
                "enrollments/admin_form.html",
                students=students,
                courses=courses,
                selected_student_id=selected_student_id,
                selected_course_id=selected_course_id,
            )

        success, message, category = _submit_enrollment(
            student=student,
            course=course,
            operator_user_id=session["user_id"],
            actor_label="管理员代学生选课",
        )
        flash(message, category)
        if success:
            return redirect(url_for("list_enrollments"))

    return render_template(
        "enrollments/admin_form.html",
        students=students,
        courses=courses,
        selected_student_id=selected_student_id,
        selected_course_id=selected_course_id,
    )


@app.route("/enrollments/withdraw/<int:enrollment_id>", methods=["POST"])
@login_required
def withdraw_enrollment(enrollment_id):
    enrollment = Enrollment.query.get_or_404(enrollment_id)

    if enrollment.status == "withdrawn":
        flash("该记录已退课", "warning")
        return redirect(url_for("list_enrollments"))

    role = session.get("role")
    if role == "student":
        flash("学生不允许退课，如需处理请联系管理员", "warning")
        return redirect(url_for("list_enrollments"))
    elif role == "teacher":
        managed_ids = _teacher_managed_class_ids(session["user_id"])
        if enrollment.student.class_id not in managed_ids:
            flash("你只能操作自己负责班级的选课记录", "danger")
            return redirect(url_for("list_enrollments"))

    try:
        enrollment.status = "withdrawn"
        enrollment.completion_date = datetime.utcnow()
        db.session.commit()

        db.session.add(OperationLog(
            user_id=session["user_id"],
            action="withdraw",
            module="enrollment",
            record_id=enrollment.id,
            details=f"退课：{enrollment.student.name}-{enrollment.course.course_name}"
        ))
        db.session.commit()

        flash("退课成功", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"退课失败：{str(e)}", "danger")

    return redirect(url_for("list_enrollments"))


# ======================== 成绩管理 ========================

@app.route("/scores")
@login_required
def list_scores():
    if session.get("role") == "student":
        student = Student.query.filter_by(user_id=session["user_id"]).first()
        if not student:
            flash("当前账号未绑定学生档案，请联系管理员处理", "warning")
            return redirect(url_for("dashboard"))
        q = Enrollment.query.filter_by(student_id=student.id)
        enrollments = q.all()
    elif session.get("role") == "teacher":
        teacher_course_ids = _teacher_course_ids(session["user_id"])
        if teacher_course_ids:
            enrollments = Enrollment.query.filter(Enrollment.course_id.in_(teacher_course_ids)).all()
        else:
            enrollments = []
    else:
        enrollments = Enrollment.query.all()

    return render_template("scores/list.html", enrollments=enrollments)


@app.route("/scores/record-manual", methods=["GET", "POST"])
@role_required("admin", "teacher")
def record_score_manual():
    students_q = Student.query
    courses_q = Course.query

    if session.get("role") == "teacher":
        teacher_course_ids = _teacher_course_ids(session["user_id"])
        if teacher_course_ids:
            courses_q = courses_q.filter(Course.id.in_(teacher_course_ids))
        else:
            courses_q = courses_q.filter(Course.id == -1)

    students = students_q.order_by(Student.student_no.asc()).all()
    courses = courses_q.order_by(Course.course_no.asc()).all()

    if request.method == "POST":
        student_id = request.form.get("student_id", type=int)
        course_id = request.form.get("course_id", type=int)
        score_value = request.form.get("score", type=float)
        grade = request.form.get("grade", "").strip()

        if not student_id or not course_id:
            flash("请选择学生和课程", "danger")
            return render_template(
                "scores/manual_form.html",
                students=students,
                courses=courses,
                selected_student_id=student_id,
                selected_course_id=course_id,
                score_value=request.form.get("score", "").strip(),
                grade=grade,
            )

        student = Student.query.get(student_id)
        course = Course.query.get(course_id)
        if not student or not course:
            flash("学生或课程不存在", "danger")
            return render_template(
                "scores/manual_form.html",
                students=students,
                courses=courses,
                selected_student_id=student_id,
                selected_course_id=course_id,
                score_value=request.form.get("score", "").strip(),
                grade=grade,
            )

        if session.get("role") == "teacher":
            teacher_course_ids = _teacher_course_ids(session["user_id"])
            if course.id not in teacher_course_ids:
                flash("你只能录入自己任课课程的成绩", "danger")
                return redirect(url_for("record_score_manual"))

        if score_value is None or score_value < 0 or score_value > 100:
            flash("成绩必须在 0-100 之间", "danger")
            return render_template(
                "scores/manual_form.html",
                students=students,
                courses=courses,
                selected_student_id=student_id,
                selected_course_id=course_id,
                score_value=request.form.get("score", "").strip(),
                grade=grade,
            )

        if not grade:
            if score_value >= 90:
                grade = "A"
            elif score_value >= 80:
                grade = "B"
            elif score_value >= 70:
                grade = "C"
            elif score_value >= 60:
                grade = "D"
            else:
                grade = "F"

        enrollment = Enrollment.query.filter_by(student_id=student_id, course_id=course_id).first()
        try:
            if not enrollment:
                enrollment = Enrollment(student_id=student_id, course_id=course_id, status="enrolled")
                db.session.add(enrollment)
                db.session.flush()

            enrollment.score = score_value
            enrollment.grade = grade
            db.session.commit()

            db.session.add(OperationLog(
                user_id=session["user_id"],
                action="record_score",
                module="score",
                record_id=enrollment.id,
                details=f"手动录入成绩：{student.name}-{course.course_name} {score_value}分"
            ))
            db.session.commit()

            flash("成绩录入成功", "success")
            return redirect(url_for("list_scores"))
        except Exception as e:
            db.session.rollback()
            flash(f"录入失败：{str(e)}", "danger")

    return render_template(
        "scores/manual_form.html",
        students=students,
        courses=courses,
        selected_student_id=None,
        selected_course_id=None,
        score_value="",
        grade="",
    )


@app.route("/scores/record/<int:enrollment_id>", methods=["GET", "POST"])
@role_required("admin", "teacher")
def record_score(enrollment_id):
    enrollment = Enrollment.query.get_or_404(enrollment_id)

    if session.get("role") == "teacher":
        teacher_course_ids = _teacher_course_ids(session["user_id"])
        if enrollment.course_id not in teacher_course_ids:
            flash("你只能录入自己任课课程的成绩", "error")
            return redirect(url_for("list_scores"))

    if request.method == "POST":
        score_value = request.form.get("score", type=float)
        grade = request.form.get("grade", "").strip()

        if score_value is None or score_value < 0 or score_value > 100:
            flash("成绩必须在 0-100 之间", "danger")
            return render_template("scores/form.html", enrollment=enrollment, action="record")

        try:
            enrollment.score = score_value
            enrollment.grade = grade

            db.session.commit()

            log = OperationLog(
                user_id=session["user_id"],
                action="record_score",
                module="score",
                record_id=enrollment.id,
                details=f"录入成绩 {score_value} 分"
            )
            db.session.add(log)
            db.session.commit()

            flash("成绩录入成功", "success")
            return redirect(url_for("list_scores"))
        except Exception as e:
            db.session.rollback()
            flash(f"录入失败：{str(e)}", "danger")

    return render_template("scores/form.html", enrollment=enrollment, action="record")


# ======================== 系统管理 ========================

@app.route("/ops-log")
@role_required("admin")
def list_ops_log():
    now = datetime.utcnow()
    key_status = request.args.get("key_status", "valid").strip().lower()
    logs = OperationLog.query.order_by(OperationLog.created_at.desc()).limit(500).all()
    reset_keys_q = (
        db.session.query(PasswordResetKey, User.username, User.full_name)
        .outerjoin(User, PasswordResetKey.user_id == User.id)
    )

    if key_status == "used":
        reset_keys_q = reset_keys_q.filter(PasswordResetKey.is_used.is_(True))
    elif key_status == "expired":
        reset_keys_q = reset_keys_q.filter(
            PasswordResetKey.is_used.is_(False),
            PasswordResetKey.expires_at < now,
        )
    elif key_status == "all":
        pass
    else:
        key_status = "valid"
        reset_keys_q = reset_keys_q.filter(
            PasswordResetKey.is_used.is_(False),
            PasswordResetKey.expires_at >= now,
        )

    reset_keys = (
        reset_keys_q
        .order_by(PasswordResetKey.created_at.desc())
        .limit(200)
        .all()
    )
    return render_template(
        "admin/logs.html",
        logs=logs,
        reset_keys=reset_keys,
        key_status=key_status,
        now=now,
    )


@app.route("/users")
@role_required("admin")
def list_users():
    users = User.query.order_by(User.id.desc()).all()
    return render_template("admin/users.html", users=users)


@app.route("/users/add", methods=["GET", "POST"])
@role_required("admin")
def add_user():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        full_name = request.form.get("full_name", "").strip()
        role = request.form.get("role", "").strip()
        email = request.form.get("email", "").strip()

        if not username or not password or not full_name:
            flash("用户名、密码、姓名不能为空", "danger")
            return render_template("admin/user_form.html", action="add")

        if User.query.filter_by(username=username).first():
            flash("用户名已存在", "danger")
            return render_template("admin/user_form.html", action="add")

        try:
            user = User(
                username=username,
                full_name=full_name,
                role=role,
                email=email
            )
            user.set_password(password)
            db.session.add(user)
            db.session.commit()

            log = OperationLog(
                user_id=session["user_id"],
                action="create",
                module="user",
                record_id=user.id,
                details=f"新增用户 {username}({full_name})"
            )
            db.session.add(log)
            db.session.commit()

            flash("新增用户成功", "success")
            return redirect(url_for("list_users"))
        except Exception as e:
            db.session.rollback()
            flash(f"添加失败：{str(e)}", "danger")

    return render_template("admin/user_form.html", user=None, action="add")


@app.route("/users/edit/<int:user_id>", methods=["GET", "POST"])
@role_required("admin")
def edit_user(user_id):
    user = User.query.get_or_404(user_id)

    if request.method == "POST":
        full_name    = request.form.get("full_name", "").strip()
        role         = request.form.get("role", "").strip()
        email        = request.form.get("email", "").strip()
        phone        = request.form.get("phone", "").strip()
        new_password = request.form.get("password", "").strip()
        is_active    = request.form.get("is_active") == "on"

        if not full_name:
            flash("姓名不能为空", "danger")
            return render_template("admin/user_form.html", user=user, action="edit")

        if new_password and len(new_password) < 6:
            flash("密码至少 6 个字符", "danger")
            return render_template("admin/user_form.html", user=user, action="edit")

        user.full_name = full_name
        user.role      = role
        user.email     = email or None
        user.phone     = phone or None
        user.is_active = is_active
        if new_password:
            user.set_password(new_password)

        try:
            db.session.commit()
            log = OperationLog(
                user_id=session["user_id"],
                action="update",
                module="user",
                record_id=user.id,
                details=f"修改用户 {user.username}({full_name})"
            )
            db.session.add(log)
            db.session.commit()
            flash("修改成功", "success")
            return redirect(url_for("list_users"))
        except Exception as e:
            db.session.rollback()
            flash(f"修改失败：{str(e)}", "danger")

    return render_template("admin/user_form.html", user=user, action="edit")


@app.route("/users/delete/<int:user_id>", methods=["POST"])
@role_required("admin")
def delete_user(user_id):
    user = User.query.get_or_404(user_id)

    if user.id == session.get("user_id"):
        flash("不能删除当前登录账号", "error")
        return redirect(url_for("list_users"))

    if user.role == "admin":
        admin_count = User.query.filter_by(role="admin").count()
        if admin_count <= 1:
            flash("系统至少需要保留一个管理员账号", "error")
            return redirect(url_for("list_users"))

    if Teacher.query.filter_by(user_id=user.id).first():
        flash("该用户绑定了教师档案，请先删除教师档案", "error")
        return redirect(url_for("list_users"))

    if Student.query.filter_by(user_id=user.id).first():
        flash("该用户绑定了学生档案，请先删除学生档案", "error")
        return redirect(url_for("list_users"))

    managed_classes = Class.query.filter_by(headteacher_id=user.id).count()
    if managed_classes > 0:
        flash("该用户仍担任辅导员，请先调整班级辅导员", "error")
        return redirect(url_for("list_users"))

    try:
        log = OperationLog(
            user_id=session["user_id"],
            action="delete",
            module="user",
            record_id=user.id,
            details=f"删除用户 {user.username}({user.full_name})"
        )
        db.session.add(log)
        db.session.delete(user)
        db.session.commit()

        flash("删除成功", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"删除失败：{str(e)}", "danger")

    return redirect(url_for("list_users"))


# ======================== 批量导入 ========================

def _make_template_workbook(headers, example_row, sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    fill = PatternFill("solid", fgColor="4F46E5")
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) * 1.8, 16)
    ws.append(example_row)
    return wb


@app.route("/students/import/template")
@role_required("admin", "teacher")
def student_import_template():
    headers = ["学号*", "姓名*", "性别", "年龄", "班级编号", "手机", "邮箱", "专业", "身份证号",
               "账号(空则用学号)", "密码(空则123456)"]
    example = ["20240001", "张三", "男", 20, "C001", "13800138000",
               "zhangsan@school.com", "计算机科学", "", "", ""]
    wb = _make_template_workbook(headers, example, "学生导入模板")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="学生导入模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/students/import", methods=["POST"])
@role_required("admin", "teacher")
def import_students():
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("请上传 .xlsx 格式的文件", "danger")
        return redirect(url_for("list_students"))

    file.stream.seek(0, 2)
    if file.stream.tell() > 5 * 1024 * 1024:
        flash("文件大小不能超过 5MB", "danger")
        return redirect(url_for("list_students"))
    file.stream.seek(0)

    try:
        wb = load_workbook(file.stream, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        flash("文件解析失败，请使用正确的模板", "danger")
        return redirect(url_for("list_students"))

    success, skipped, errors = 0, 0, []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v for v in row if v not in (None, "")):
            continue

        def _s(v): return str(v).strip() if v is not None else ""
        student_no = _s(row[0])
        name       = _s(row[1])
        gender     = _s(row[2])
        age_raw    = row[3]
        class_no   = _s(row[4])
        phone      = _s(row[5])
        email      = _s(row[6])
        major      = _s(row[7])
        id_number  = _s(row[8])
        username   = _s(row[9]) or student_no
        password   = _s(row[10]) or "123456"

        if not student_no or not name:
            errors.append(f"第 {row_idx} 行：学号和姓名不能为空，已跳过")
            skipped += 1
            continue
        if Student.query.filter_by(student_no=student_no).first():
            errors.append(f"第 {row_idx} 行：学号 {student_no} 已存在，已跳过")
            skipped += 1
            continue
        if User.query.filter_by(username=username).first():
            errors.append(f"第 {row_idx} 行：账号 {username} 已被占用，已跳过")
            skipped += 1
            continue

        class_obj = Class.query.filter_by(class_no=class_no).first() if class_no else None
        try:
            age = int(float(str(age_raw))) if age_raw not in (None, "") else None
        except (ValueError, TypeError):
            age = None

        try:
            user = User(username=username, full_name=name, role="student",
                        email=email or None, phone=phone or None)
            user.set_password(password)
            db.session.add(user)
            db.session.flush()
            student = Student(
                user_id=user.id,
                student_no=student_no,
                name=name, gender=gender, age=age,
                class_id=class_obj.id if class_obj else None,
                phone=phone, email=email, major=major,
                id_number=id_number or None,
                status="enrolled"
            )
            db.session.add(student)
            db.session.commit()
            success += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"第 {row_idx} 行：{name}({student_no}) 导入失败 — {e}")
            skipped += 1

    db.session.add(OperationLog(
        user_id=session["user_id"], action="batch_import", module="student",
        details=f"批量导入学生：成功 {success} 条，跳过 {skipped} 条"
    ))
    db.session.commit()

    for err in errors[:5]:
        flash(err, "warning")
    if len(errors) > 5:
        flash(f"……还有 {len(errors) - 5} 条错误未显示", "warning")
    flash(f"导入完成：成功 {success} 条，跳过 {skipped} 条",
          "success" if success > 0 else "warning")
    return redirect(url_for("list_students"))


@app.route("/teachers/import/template")
@role_required("admin")
def teacher_import_template():
    headers = ["教工号*", "姓名*", "性别", "部门", "手机", "邮箱", "学历",
               "账号(空则用教工号)", "密码(空则123456)"]
    example = ["T20240001", "李四", "女", "计算机学院", "13900139000",
               "lisi@school.com", "博士", "", ""]
    wb = _make_template_workbook(headers, example, "教师导入模板")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="教师导入模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/teachers/import", methods=["POST"])
@role_required("admin")
def import_teachers():
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("请上传 .xlsx 格式的文件", "danger")
        return redirect(url_for("list_teachers"))

    file.stream.seek(0, 2)
    if file.stream.tell() > 5 * 1024 * 1024:
        flash("文件大小不能超过 5MB", "danger")
        return redirect(url_for("list_teachers"))
    file.stream.seek(0)

    try:
        wb = load_workbook(file.stream, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        flash("文件解析失败，请使用正确的模板", "danger")
        return redirect(url_for("list_teachers"))

    success, skipped, errors = 0, 0, []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v for v in row if v not in (None, "")):
            continue

        def _s(v): return str(v).strip() if v is not None else ""
        teacher_no    = _s(row[0])
        name          = _s(row[1])
        gender        = _s(row[2])
        department    = _s(row[3])
        phone         = _s(row[4])
        email         = _s(row[5])
        qualification = _s(row[6])
        username      = _s(row[7]) or teacher_no
        password      = _s(row[8]) or "123456"

        if not teacher_no or not name:
            errors.append(f"第 {row_idx} 行：教工号和姓名不能为空，已跳过")
            skipped += 1
            continue
        if Teacher.query.filter_by(teacher_no=teacher_no).first():
            errors.append(f"第 {row_idx} 行：教工号 {teacher_no} 已存在，已跳过")
            skipped += 1
            continue
        if User.query.filter_by(username=username).first():
            errors.append(f"第 {row_idx} 行：账号 {username} 已被占用，已跳过")
            skipped += 1
            continue

        try:
            user = User(username=username, full_name=name, role="teacher",
                        email=email or None, phone=phone or None)
            user.set_password(password)
            db.session.add(user)
            db.session.flush()
            teacher = Teacher(
                user_id=user.id,
                teacher_no=teacher_no,
                name=name, gender=gender,
                department=department, phone=phone,
                email=email, qualification=qualification
            )
            db.session.add(teacher)
            db.session.commit()
            success += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"第 {row_idx} 行：{name}({teacher_no}) 导入失败 — {e}")
            skipped += 1

    db.session.add(OperationLog(
        user_id=session["user_id"], action="batch_import", module="teacher",
        details=f"批量导入教师：成功 {success} 条，跳过 {skipped} 条"
    ))
    db.session.commit()

    for err in errors[:5]:
        flash(err, "warning")
    if len(errors) > 5:
        flash(f"……还有 {len(errors) - 5} 条错误未显示", "warning")
    flash(f"导入完成：成功 {success} 条，跳过 {skipped} 条",
          "success" if success > 0 else "warning")
    return redirect(url_for("list_teachers"))


@app.route("/courses/import/template")
@role_required("admin")
def course_import_template():
    headers = ["课程号*", "课程名称*", "学分", "学时", "任课教师工号", "容量", "学期", "状态(open/closed)"]
    example = ["CS101", "程序设计基础", 3, 48, "T20240001", 60, "2026-1", "open"]
    wb = _make_template_workbook(headers, example, "课程导入模板")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="课程导入模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/courses/import", methods=["POST"])
@role_required("admin")
def import_courses():
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("请上传 .xlsx 格式的文件", "danger")
        return redirect(url_for("list_courses"))

    file.stream.seek(0, 2)
    if file.stream.tell() > 5 * 1024 * 1024:
        flash("文件大小不能超过 5MB", "danger")
        return redirect(url_for("list_courses"))
    file.stream.seek(0)

    try:
        wb = load_workbook(file.stream, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        flash("文件解析失败，请使用正确的模板", "danger")
        return redirect(url_for("list_courses"))

    success, skipped, errors = 0, 0, []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v for v in row if v not in (None, "")):
            continue

        def _s(v): return str(v).strip() if v is not None else ""

        course_no = _s(row[0])
        course_name = _s(row[1])
        credits_raw = row[2]
        hours_raw = row[3]
        teacher_no = _s(row[4])
        capacity_raw = row[5]
        semester = _s(row[6])
        status = (_s(row[7]).lower() or "open")

        if not course_no or not course_name:
            errors.append(f"第 {row_idx} 行：课程号和课程名称不能为空，已跳过")
            skipped += 1
            continue
        if Course.query.filter_by(course_no=course_no).first():
            errors.append(f"第 {row_idx} 行：课程号 {course_no} 已存在，已跳过")
            skipped += 1
            continue

        try:
            credits = float(credits_raw) if credits_raw not in (None, "") else None
        except (TypeError, ValueError):
            credits = None

        try:
            hours = int(float(str(hours_raw))) if hours_raw not in (None, "") else None
        except (TypeError, ValueError):
            hours = None

        try:
            max_capacity = int(float(str(capacity_raw))) if capacity_raw not in (None, "") else 50
        except (TypeError, ValueError):
            max_capacity = 50

        if status not in ["open", "closed"]:
            status = "open"

        teacher_id = None
        if teacher_no:
            teacher_obj = Teacher.query.filter_by(teacher_no=teacher_no).first()
            if not teacher_obj:
                errors.append(f"第 {row_idx} 行：任课教师工号 {teacher_no} 不存在，已跳过")
                skipped += 1
                continue
            teacher_id = teacher_obj.id

        try:
            course = Course(
                course_no=course_no,
                course_name=course_name,
                credits=credits,
                hours=hours,
                teacher_id=teacher_id,
                max_capacity=max_capacity if max_capacity > 0 else 50,
                semester=semester,
                status=status,
            )
            db.session.add(course)
            db.session.commit()
            success += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"第 {row_idx} 行：{course_name}({course_no}) 导入失败 — {e}")
            skipped += 1

    db.session.add(OperationLog(
        user_id=session["user_id"], action="batch_import", module="course",
        details=f"批量导入课程：成功 {success} 条，跳过 {skipped} 条"
    ))
    db.session.commit()

    for err in errors[:5]:
        flash(err, "warning")
    if len(errors) > 5:
        flash(f"……还有 {len(errors) - 5} 条错误未显示", "warning")
    flash(f"导入完成：成功 {success} 条，跳过 {skipped} 条",
          "success" if success > 0 else "warning")
    return redirect(url_for("list_courses"))


# ======================== 管理员路由 ========================

@app.route("/admin/")
@login_required
@role_required("admin")
def admin_index():
    """管理员首页"""
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    backups = sorted([f for f in os.listdir(backup_dir) if f.endswith(".db")], reverse=True)
    backup_count = len(backups)
    latest_backup = backups[0] if backups else None

    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
    log_sizes = {}
    if os.path.exists(log_dir):
        for f in os.listdir(log_dir):
            if f.endswith(".log"):
                log_sizes[f] = os.path.getsize(os.path.join(log_dir, f)) / 1024

    stats = {
        "backup_count": backup_count,
        "latest_backup": latest_backup,
        "log_sizes": log_sizes,
    }
    return render_template("admin/index.html", stats=stats)


@app.route("/admin/backup")
@login_required
@role_required("admin")
def admin_backup_list():
    """备份列表"""
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    backups = []
    for f in sorted(os.listdir(backup_dir), reverse=True):
        if f.endswith(".db"):
            fpath = os.path.join(backup_dir, f)
            size_mb = round(os.path.getsize(fpath) / (1024 * 1024), 2)
            mtime = datetime.fromtimestamp(os.path.getmtime(fpath)).strftime("%Y-%m-%d %H:%M:%S")
            backups.append({"filename": f, "size_mb": size_mb, "date": mtime})

    return render_template("admin/backup.html", backups=backups)


@app.route("/admin/backup/create", methods=["POST"])
@login_required
@role_required("admin")
def admin_create_backup():
    """创建备份"""
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "instance", "students.db")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"backup_{timestamp}.db"
    backup_path = os.path.join(backup_dir, backup_name)

    try:
        shutil.copy2(db_path, backup_path)
        db.session.add(OperationLog(
            user_id=session["user_id"], action="create_backup", module="admin",
            details=f"创建备份: {backup_name}"
        ))
        db.session.commit()
        flash(f"备份创建成功: {backup_name}", "success")
    except Exception as e:
        flash(f"备份失败: {str(e)}", "danger")

    return redirect(url_for("admin_backup_list"))


@app.route("/admin/backup/restore/<backup_file>", methods=["POST"])
@login_required
@role_required("admin")
def admin_restore_backup(backup_file):
    """恢复备份"""
    if not backup_file.endswith(".db") or "/" in backup_file or "\\" in backup_file:
        flash("无效的备份文件名", "danger")
        return redirect(url_for("admin_backup_list"))

    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    backup_path = os.path.join(backup_dir, backup_file)
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "instance", "students.db")

    if not os.path.exists(backup_path):
        flash("备份文件不存在", "danger")
        return redirect(url_for("admin_backup_list"))

    try:
        shutil.copy2(backup_path, db_path)
        db.session.add(OperationLog(
            user_id=session["user_id"], action="restore_backup", module="admin",
            details=f"恢复备份: {backup_file}"
        ))
        db.session.commit()
        flash("备份恢复成功，请重启应用以确保数据完全生效", "success")
    except Exception as e:
        flash(f"恢复失败: {str(e)}", "danger")

    return redirect(url_for("admin_backup_list"))


@app.route("/admin/backup/delete/<backup_file>", methods=["POST"])
@login_required
@role_required("admin")
def admin_delete_backup(backup_file):
    """删除备份"""
    if not backup_file.endswith(".db") or "/" in backup_file or "\\" in backup_file:
        flash("无效的备份文件名", "danger")
        return redirect(url_for("admin_backup_list"))

    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    backup_path = os.path.join(backup_dir, backup_file)

    if not os.path.exists(backup_path):
        flash("备份文件不存在", "danger")
        return redirect(url_for("admin_backup_list"))

    try:
        os.remove(backup_path)
        db.session.add(OperationLog(
            user_id=session["user_id"], action="delete_backup", module="admin",
            details=f"删除备份: {backup_file}"
        ))
        db.session.commit()
        flash(f"备份已删除: {backup_file}", "success")
    except Exception as e:
        flash(f"删除失败: {str(e)}", "danger")

    return redirect(url_for("admin_backup_list"))


@app.route("/admin/backup/download/<backup_file>")
@login_required
@role_required("admin")
def admin_download_backup(backup_file):
    """下载备份"""
    if not backup_file.endswith(".db") or "/" in backup_file or "\\" in backup_file:
        flash("无效的备份文件名", "danger")
        return redirect(url_for("admin_backup_list"))

    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    backup_path = os.path.join(backup_dir, backup_file)

    if not os.path.exists(backup_path):
        flash("备份文件不存在", "danger")
        return redirect(url_for("admin_backup_list"))

    return send_file(backup_path, as_attachment=True, download_name=backup_file)


@app.route("/admin/data")
@login_required
@role_required("admin")
def admin_data():
    """数据管理页"""
    return render_template("admin/data.html")


@app.route("/admin/data/export/students", methods=["POST"])
@login_required
@role_required("admin")
def admin_export_students():
    """导出学生数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息"
    ws.append(["学号", "姓名", "性别", "班级", "电话", "邮箱", "身份证号", "户籍地址", "现居住地", "入学日期", "状态", "备注"])

    for s in Student.query.all():
        ws.append([
            s.student_no, s.name, s.gender,
            s.class_obj.class_name if s.class_obj else "",
            s.phone or "", s.email or "", s.id_number or "",
            s.home_address or "", s.current_address or "",
            str(s.enrollment_date) if s.enrollment_date else "",
            s.status or "", s.notes or ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/data/export/scores", methods=["POST"])
@login_required
@role_required("admin")
def admin_export_scores():
    """导出成绩数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩数据"
    ws.append(["学号", "学生姓名", "课程名称", "成绩", "等级", "学期", "录入时间"])

    for s in Score.query.join(Enrollment).join(Student).join(Course).all():
        enrollment = Enrollment.query.get(s.enrollment_id)
        student = enrollment.student if enrollment else None
        course = enrollment.course if enrollment else None
        ws.append([
            student.student_no if student else "",
            student.name if student else "",
            course.course_name if course else "",
            s.score_value if s.score_value is not None else "",
            s.grade or "",
            course.semester if course else "",
            str(s.recorded_at) if s.recorded_at else ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"scores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/data/export/courses", methods=["POST"])
@login_required
@role_required("admin")
def admin_export_courses():
    """导出课程数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "课程数据"
    ws.append(["课程编号", "课程名称", "学分", "课时", "授课教师", "学期", "状态"])

    for c in Course.query.all():
        ws.append([
            c.course_no, c.course_name,
            c.credits or "", c.hours or "",
            c.teacher.name if c.teacher else "",
            c.semester or "", c.status or ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"courses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/cleanup-logs", methods=["POST"])
@login_required
@role_required("admin")
def admin_cleanup_logs():
    """清空当前所有日志文件"""
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
    cleaned = 0
    failed = 0

    if os.path.exists(log_dir):
        for f in os.listdir(log_dir):
            if f.endswith(".log") or ".log." in f:
                fpath = os.path.join(log_dir, f)
                if not os.path.isfile(fpath):
                    continue
                try:
                    # 清空文件而非删除，避免运行中的日志句柄失效。
                    with open(fpath, "w", encoding="utf-8"):
                        pass
                    cleaned += 1
                except Exception:
                    failed += 1

    db.session.add(OperationLog(
        user_id=session["user_id"], action="cleanup_logs", module="admin",
        details=f"清理日志: 清空 {cleaned} 个文件, 失败 {failed} 个"
    ))
    db.session.commit()
    if failed:
        flash(f"日志清理完成：清空 {cleaned} 个文件，{failed} 个文件清理失败", "warning")
    else:
        flash(f"日志清理完成：已清空 {cleaned} 个日志文件", "success")
    return redirect(url_for("admin_index"))


@app.route("/admin/system-info")
@login_required
@role_required("admin")
def admin_system_info():
    """系统信息"""
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")

    backup_size = (
        sum(os.path.getsize(os.path.join(backup_dir, f))
            for f in os.listdir(backup_dir)
            if os.path.isfile(os.path.join(backup_dir, f)))
        / (1024 * 1024)
    ) if os.path.exists(backup_dir) else 0

    log_size = (
        sum(os.path.getsize(os.path.join(log_dir, f))
            for f in os.listdir(log_dir)
            if os.path.isfile(os.path.join(log_dir, f)))
        / 1024
    ) if os.path.exists(log_dir) else 0

    info = {
        "python_version": platform.python_version(),
        "platform": platform.platform(),
        "processor": platform.processor(),
        "backup_dir_size": backup_size,
        "log_dir_size": log_size,
    }
    return render_template("admin/system_info.html", info=info)


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5001)
